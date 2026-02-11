"""
Callbacks del detalle de cliente.
Carga informacion del cliente y construye arbol de ventas: generico -> marca -> articulo.
Muestra bultos por mes en columnas. Articulos sin venta aparecen con 0.
"""
import io
from dash import callback, Output, Input, State, html, dcc, no_update, clientside_callback
import pandas as pd

# openpyxl se importa lazy (solo al exportar Excel) para no ralentizar el startup
_openpyxl_styles = None


def _get_excel_styles():
    """Importa openpyxl y crea estilos la primera vez que se necesitan."""
    global _openpyxl_styles
    if _openpyxl_styles is None:
        from openpyxl.styles import Font, PatternFill, Alignment
        _openpyxl_styles = {
            'FONT_BOLD': Font(bold=True),
            'FONT_HEADER': Font(bold=True, size=11),
            'FONT_GENERICO': Font(bold=True, size=12),
            'FONT_MARCA': Font(bold=True, size=11, color='333333'),
            'FONT_TOTAL': Font(bold=True, size=11),
            'FONT_CLIENTE': Font(bold=True, size=14),
            'FILL_HEADER': PatternFill('solid', fgColor='E0E0E0'),
            'FILL_GENERICO': PatternFill('solid', fgColor='D6E4F0'),
            'FILL_MARCA': PatternFill('solid', fgColor='F0F0F0'),
            'FILL_SUBTOTAL': PatternFill('solid', fgColor='F5F8FC'),
            'FILL_TOTAL': PatternFill('solid', fgColor='D6E4F0'),
            'ALIGN_RIGHT': Alignment(horizontal='right'),
            'ALIGN_CENTER': Alignment(horizontal='center'),
        }
    return _openpyxl_styles

from data.queries import cargar_info_cliente, cargar_ventas_cliente_detalle

MESES_CORTOS = {
    1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
    7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
}


@callback(
    [Output('cliente-header-content', 'children'),
     Output('cliente-detail-content', 'children')],
    [Input('cliente-id-store', 'data')]
)
def cargar_detalle_cliente(store_data):
    """Carga datos del cliente y construye la vista de detalle."""
    if not store_data:
        return html.Div(), html.Div("No se encontro el cliente")

    id_cliente = store_data['id_cliente']

    # Cargar datos (2 queries: info cliente + todos los articulos con/sin venta)
    df_info = cargar_info_cliente(id_cliente)
    df_all = cargar_ventas_cliente_detalle(id_cliente)

    # === HEADER ===
    if len(df_info) > 0:
        info = df_info.iloc[0]
        header_children = [
            html.H1([
                info['razon_social'],
                html.Span(f"  [{id_cliente}]", style={'fontSize': '16px', 'color': '#999', 'fontWeight': 'normal'}),
            ], style={
                'color': 'white', 'margin': '0', 'fontSize': '28px'
            }),
        ]
        if info.get('fantasia'):
            header_children.append(
                html.P(info['fantasia'], style={'color': '#ccc', 'margin': '5px 0 0 0', 'fontSize': '16px'})
            )
        # Badges de info
        badges = []
        for label, value in [
            ('Localidad', info['localidad']),
            ('Canal', info['canal']),
            ('Sucursal', info['sucursal']),
            ('Ramo', info['ramo']),
            ('Lista Precio', info['lista_precio']),
        ]:
            badges.append(html.Div([
                html.Span(f"{label}: ", style={'color': '#999', 'fontSize': '13px'}),
                html.Span(str(value), style={'color': 'white', 'fontSize': '13px', 'fontWeight': 'bold'}),
            ]))
        header_children.append(
            html.Div(badges, style={'display': 'flex', 'gap': '20px', 'marginTop': '12px', 'flexWrap': 'wrap'})
        )
        header = html.Div(header_children)
    else:
        header = html.H1(f"Cliente {id_cliente}", style={'color': 'white', 'margin': '0'})

    # === CONTENIDO ===
    if len(df_all) == 0:
        content = html.Div(
            "No hay articulos disponibles.",
            style={'textAlign': 'center', 'color': '#666', 'padding': '60px', 'fontSize': '18px'}
        )
        return header, content

    # Separar articulos con y sin venta (bultos NULL = sin venta)
    df_con_venta = df_all.dropna(subset=['bultos'])
    df_con_venta = df_con_venta.copy()
    df_con_venta['bultos'] = df_con_venta['bultos'].astype(float)

    # Articulos sin venta: filas con bultos NULL (sin ningun periodo)
    arts_con_venta = set(df_con_venta['id_articulo'].unique())
    df_sin_venta_arts = df_all[~df_all['id_articulo'].isin(arts_con_venta)].drop_duplicates(subset=['id_articulo'])

    # Periodos solo de articulos con venta (ultimos 12 meses)
    periodos = sorted(
        df_con_venta[['anio', 'mes']].drop_duplicates().values.tolist()
    ) if len(df_con_venta) > 0 else []
    periodos = periodos[-12:]

    # Rearmar df_all: con_venta + sin_venta con bultos=0
    if len(df_sin_venta_arts) > 0:
        df_sv = df_sin_venta_arts[['id_articulo', 'generico', 'marca', 'articulo']].copy()
        df_sv['anio'] = pd.NA
        df_sv['mes'] = pd.NA
        df_sv['bultos'] = 0.0
        df_all = pd.concat([df_con_venta, df_sv], ignore_index=True)
    else:
        df_all = df_con_venta

    # KPIs del mes actual
    from datetime import date
    hoy = date.today()
    mes_actual_anio = hoy.year
    mes_actual_mes = hoy.month
    mes_label = _periodo_label(mes_actual_anio, mes_actual_mes)

    df_mes = df_con_venta[(df_con_venta['anio'] == mes_actual_anio) & (df_con_venta['mes'] == mes_actual_mes)]
    mes_bultos = df_mes['bultos'].sum() if len(df_mes) > 0 else 0
    mes_genericos = df_mes['generico'].nunique() if len(df_mes) > 0 else 0
    mes_articulos = df_mes['articulo'].nunique() if len(df_mes) > 0 else 0

    n_con_venta = df_con_venta['id_articulo'].nunique() if len(df_con_venta) > 0 else 0
    n_sin_venta = len(df_sin_venta_arts)

    ayuda_kpis = html.Div(
        f"Los primeros 3 indicadores corresponden al mes actual ({mes_label}). "
        f"\"Con Venta\" y \"Sin Venta\" reflejan el historico completo del catalogo.",
        style={'fontSize': '12px', 'color': '#999', 'marginBottom': '10px'}
    )

    resumen = html.Div([
        ayuda_kpis,
        html.Div([
            _resumen_card(f"Bultos {mes_label}", f"{mes_bultos:,.0f}"),
            _resumen_card(f"Categorias {mes_label}", f"{mes_genericos}"),
            _resumen_card(f"Articulos {mes_label}", f"{mes_articulos}"),
            _resumen_card("Con Venta", f"{n_con_venta}"),
            _resumen_card("Sin Venta", f"{n_sin_venta}"),
        ], style={
            'display': 'flex', 'gap': '20px', 'flexWrap': 'wrap'
        }),
    ], style={'marginBottom': '25px'})

    # Tabla plana con subtotales
    tabla = _build_flat_table(df_all, periodos)

    # Dropdowns de salto rápido
    gen_totals_sorted = df_all.groupby('generico')['bultos'].sum().sort_values(ascending=False)
    generico_options = [{'label': g, 'value': f"gen-{g.replace(' ', '-').lower()}"} for g in gen_totals_sorted.index]

    marca_options = []
    for generico in gen_totals_sorted.index:
        df_gen = df_all[df_all['generico'] == generico]
        marcas_sorted = df_gen.groupby('marca')['bultos'].sum().sort_values(ascending=False)
        for marca in marcas_sorted.index:
            marca_options.append({
                'label': f"{generico} > {marca}",
                'value': f"marca-{generico.replace(' ', '-').lower()}-{marca.replace(' ', '-').lower()}"
            })

    jump_bar = html.Div([
        html.Div([
            dcc.Dropdown(
                id='jump-generico',
                options=generico_options,
                placeholder='Ir a generico...',
                style={'fontSize': '13px'},
                clearable=True,
            ),
        ], style={'flex': '1', 'minWidth': '200px'}),
        html.Div([
            dcc.Dropdown(
                id='jump-marca',
                options=marca_options,
                placeholder='Ir a marca...',
                style={'fontSize': '13px'},
                clearable=True,
            ),
        ], style={'flex': '1', 'minWidth': '200px'}),
        html.Button(
            "Exportar Todo a Excel",
            id='btn-excel-completo',
            style={
                'padding': '8px 20px', 'fontSize': '14px', 'cursor': 'pointer',
                'border': '1px solid #217346', 'borderRadius': '6px',
                'backgroundColor': '#fff', 'color': '#217346', 'fontWeight': 'bold',
                'whiteSpace': 'nowrap',
            }
        ),
    ], style={
        'display': 'flex', 'gap': '15px', 'marginBottom': '15px',
        'alignItems': 'center', 'flexWrap': 'wrap',
    })

    content = html.Div([resumen, jump_bar, tabla])
    return header, content


def _resumen_card(label, value):
    """Crea una card de resumen."""
    return html.Div([
        html.Div(value, style={
            'fontSize': '28px', 'fontWeight': 'bold', 'color': '#1a1a2e'
        }),
        html.Div(label, style={'fontSize': '14px', 'color': '#666'}),
    ], style={
        'backgroundColor': 'white',
        'padding': '20px 30px',
        'borderRadius': '8px',
        'boxShadow': '0 2px 4px rgba(0,0,0,0.1)',
        'flex': '1',
        'minWidth': '150px',
    })


def _periodo_label(anio, mes):
    """Formatea un periodo como 'Ene 25'."""
    return f"{MESES_CORTOS[mes]} {str(anio)[2:]}"


def _build_flat_table(df_all, periodos):
    """Construye una tabla plana con filas de header por genérico/marca y subtotales."""
    # Estilos
    th_style = {
        'padding': '8px 10px', 'backgroundColor': '#f0f0f0',
        'textAlign': 'center', 'fontSize': '12px', 'fontWeight': 'bold',
        'borderBottom': '2px solid #ddd', 'whiteSpace': 'nowrap',
        'position': 'sticky', 'top': '0', 'zIndex': '2',
    }
    th_left = {**th_style, 'textAlign': 'left', 'minWidth': '180px'}
    th_cod = {**th_style, 'textAlign': 'center', 'minWidth': '60px'}
    th_total = {**th_style, 'backgroundColor': '#e0e7ef', 'minWidth': '70px'}
    th_excel = {**th_style, 'minWidth': '50px', 'width': '50px'}

    td_style = {
        'padding': '6px 10px', 'borderBottom': '1px solid #eee', 'fontSize': '12px',
        'textAlign': 'right', 'whiteSpace': 'nowrap',
    }
    td_left = {
        'padding': '6px 10px', 'borderBottom': '1px solid #eee', 'fontSize': '12px',
        'textAlign': 'left',
    }
    td_zero = {**td_style, 'color': '#ccc'}
    td_total = {**td_style, 'fontWeight': 'bold', 'backgroundColor': '#f5f8fc'}
    td_total_zero = {**td_total, 'color': '#ccc'}

    # Estilos de filas de grupo
    generico_row_style = {
        'backgroundColor': '#d6e4f0',
    }
    generico_cell = {
        'padding': '10px 10px', 'fontSize': '14px', 'fontWeight': 'bold',
        'borderBottom': '2px solid #b8cfe0',
    }
    marca_row_style = {
        'backgroundColor': '#eef3f8',
    }
    marca_cell = {
        'padding': '8px 10px', 'fontSize': '13px', 'fontWeight': 'bold',
        'borderBottom': '1px solid #d0d8e0', 'color': '#333',
    }
    subtotal_marca_style = {
        'backgroundColor': '#f5f8fc',
    }
    subtotal_gen_style = {
        'backgroundColor': '#e4ecf5',
    }
    subtotal_cell = {
        'padding': '6px 10px', 'fontSize': '12px', 'fontWeight': 'bold',
        'textAlign': 'right', 'borderBottom': '1px solid #ccc', 'whiteSpace': 'nowrap',
    }
    subtotal_gen_cell = {
        'padding': '8px 10px', 'fontSize': '13px', 'fontWeight': 'bold',
        'textAlign': 'right', 'borderBottom': '2px solid #b8cfe0', 'whiteSpace': 'nowrap',
    }

    n_cols = 3 + len(periodos) + 1  # cod + articulo + meses + total + excel

    # Header
    header_cells = [html.Th("Cod", style=th_cod), html.Th("Articulo", style=th_left)]
    for anio, mes in periodos:
        header_cells.append(html.Th(_periodo_label(anio, mes), style=th_style))
    header_cells.append(html.Th("Total", style=th_total))
    header_cells.append(html.Th("", style=th_excel))
    header = html.Tr(header_cells)

    rows = []

    gen_totals = df_all.groupby('generico')['bultos'].sum().sort_values(ascending=False)

    for generico in gen_totals.index:
        df_gen = df_all[df_all['generico'] == generico]
        gen_bultos = gen_totals[generico]

        # Fila header genérico
        gen_id = f"gen-{generico.replace(' ', '-').lower()}"
        rows.append(html.Tr([
            html.Td(
                generico,
                colSpan=2 + len(periodos),
                style=generico_cell,
            ),
            html.Td(f"{gen_bultos:,.0f}", style={**generico_cell, 'textAlign': 'right'}),
            html.Td("", style=generico_cell),
        ], style=generico_row_style, id=gen_id))

        marca_totals = df_gen.groupby('marca')['bultos'].sum().sort_values(ascending=False)
        gen_mes_totals = {}

        for marca in marca_totals.index:
            df_marca = df_gen[df_gen['marca'] == marca]
            marca_bultos = marca_totals[marca]

            # Fila header marca
            btn_excel = html.Button("Excel", className='btn-excel-marca',
                        **{'data-generico': generico, 'data-marca': marca},
                        style={
                            'padding': '2px 8px', 'fontSize': '11px', 'cursor': 'pointer',
                            'border': '1px solid #217346', 'borderRadius': '3px',
                            'backgroundColor': '#fff', 'color': '#217346',
                        })
            marca_id = f"marca-{generico.replace(' ', '-').lower()}-{marca.replace(' ', '-').lower()}"
            rows.append(html.Tr([
                html.Td(
                    marca,
                    colSpan=2 + len(periodos),
                    style=marca_cell,
                ),
                html.Td(f"{marca_bultos:,.0f}", style={**marca_cell, 'textAlign': 'right'}),
                html.Td(btn_excel, style={**marca_cell, 'textAlign': 'center'}),
            ], style=marca_row_style, id=marca_id))

            # Mapa de id_articulo por nombre
            art_ids = df_marca.groupby('articulo')['id_articulo'].first().to_dict()

            # Pivot de datos
            df_con_periodos = df_marca.dropna(subset=['anio', 'mes'])
            pivot = df_con_periodos.groupby(['articulo', 'anio', 'mes'])['bultos'].sum().to_dict() if len(df_con_periodos) > 0 else {}
            art_totals_marca = df_marca.groupby('articulo')['bultos'].sum().sort_values(ascending=False)

            # Filas de artículos
            for articulo in art_totals_marca.index:
                total = art_totals_marca[articulo]
                is_zero = total == 0
                cod = art_ids.get(articulo, '')
                td_cod = {
                    'padding': '6px 10px', 'borderBottom': '1px solid #eee', 'fontSize': '11px',
                    'textAlign': 'center', 'color': '#ccc' if is_zero else '#999',
                }
                cells = [
                    html.Td(str(cod), style=td_cod),
                    html.Td(articulo, style={**td_left, 'color': '#bbb', 'paddingLeft': '20px'} if is_zero else {**td_left, 'paddingLeft': '20px'}),
                ]
                for anio, mes in periodos:
                    val = pivot.get((articulo, anio, mes))
                    if val is not None:
                        cells.append(html.Td(f"{val:,.0f}", style=td_zero if val == 0 else td_style))
                    else:
                        cells.append(html.Td("", style=td_style))
                cells.append(html.Td(f"{total:,.0f}", style=td_total_zero if is_zero else td_total))
                cells.append(html.Td("", style={'borderBottom': '1px solid #eee'}))
                rows.append(html.Tr(cells))

            # Subtotal marca
            if len(df_con_periodos) > 0:
                sub_cells = [
                    html.Td("", style=subtotal_cell),
                    html.Td(f"Subtotal {marca}", style={**subtotal_cell, 'textAlign': 'left', 'paddingLeft': '20px'}),
                ]
                mes_totals = df_con_periodos.groupby(['anio', 'mes'])['bultos'].sum().to_dict()
                for anio, mes in periodos:
                    val = mes_totals.get((anio, mes))
                    sub_cells.append(html.Td(f"{val:,.0f}" if val else "", style=subtotal_cell))
                    # Acumular para subtotal genérico
                    if val:
                        gen_mes_totals[(anio, mes)] = gen_mes_totals.get((anio, mes), 0) + val
                sub_cells.append(html.Td(f"{marca_bultos:,.0f}", style={**subtotal_cell, 'borderTop': '1px solid #999'}))
                sub_cells.append(html.Td("", style=subtotal_cell))
                rows.append(html.Tr(sub_cells, style=subtotal_marca_style))

        # Subtotal genérico
        sub_gen_cells = [
            html.Td("", style=subtotal_gen_cell),
            html.Td(f"TOTAL {generico}", style={**subtotal_gen_cell, 'textAlign': 'left'}),
        ]
        for anio, mes in periodos:
            val = gen_mes_totals.get((anio, mes))
            sub_gen_cells.append(html.Td(f"{val:,.0f}" if val else "", style=subtotal_gen_cell))
        sub_gen_cells.append(html.Td(f"{gen_bultos:,.0f}", style={**subtotal_gen_cell, 'borderTop': '2px solid #666'}))
        sub_gen_cells.append(html.Td("", style=subtotal_gen_cell))
        rows.append(html.Tr(sub_gen_cells, style=subtotal_gen_style))

    table = html.Table(
        [html.Thead(header), html.Tbody(rows)],
        style={'width': '100%', 'borderCollapse': 'collapse', 'backgroundColor': 'white'}
    )

    return html.Div(
        table,
        style={
            'overflowX': 'auto', 'maxWidth': '100%',
            'maxHeight': '70vh', 'overflowY': 'auto',
            'borderRadius': '8px', 'boxShadow': '0 2px 4px rgba(0,0,0,0.1)',
        }
    )


# =============================================================================
# CLIENTSIDE CALLBACKS - SCROLL A GENÉRICO/MARCA
# =============================================================================

clientside_callback(
    """
    function(value) {
        if (!value) return window.dash_clientside.no_update;
        var el = document.getElementById(value);
        if (el) el.scrollIntoView({behavior: 'smooth', block: 'center'});
        return '';
    }
    """,
    Output('jump-generico', 'value'),
    Input('jump-generico', 'value'),
    prevent_initial_call=True
)

clientside_callback(
    """
    function(value) {
        if (!value) return window.dash_clientside.no_update;
        var el = document.getElementById(value);
        if (el) el.scrollIntoView({behavior: 'smooth', block: 'center'});
        return '';
    }
    """,
    Output('jump-marca', 'value'),
    Input('jump-marca', 'value'),
    prevent_initial_call=True
)


# =============================================================================
# HELPERS EXCEL
# =============================================================================


def _preparar_datos_excel(id_cliente):
    """Carga y prepara datos del cliente para exportar a Excel."""
    df_all = cargar_ventas_cliente_detalle(id_cliente)
    if len(df_all) == 0:
        return df_all, []

    df_con_venta = df_all.dropna(subset=['bultos']).copy()
    df_con_venta['bultos'] = df_con_venta['bultos'].astype(float)

    arts_con_venta = set(df_con_venta['id_articulo'].unique())
    df_sin_venta_arts = df_all[~df_all['id_articulo'].isin(arts_con_venta)].drop_duplicates(subset=['id_articulo'])

    periodos = sorted(
        df_con_venta[['anio', 'mes']].drop_duplicates().values.tolist()
    ) if len(df_con_venta) > 0 else []
    periodos = periodos[-12:]

    if len(df_sin_venta_arts) > 0:
        df_sv = df_sin_venta_arts[['id_articulo', 'generico', 'marca', 'articulo']].copy()
        df_sv['anio'] = pd.NA
        df_sv['mes'] = pd.NA
        df_sv['bultos'] = 0.0
        df_all = pd.concat([df_con_venta, df_sv], ignore_index=True)
    else:
        df_all = df_con_venta

    return df_all, periodos


def _escribir_tabla_marca(ws, row, df_marca, periodos, escribir_header=True):
    """Escribe una tabla de articulos por marca en la hoja. Retorna la fila siguiente."""
    from openpyxl.styles import Alignment
    s = _get_excel_styles()

    df_con_periodos = df_marca.dropna(subset=['anio', 'mes'])
    pivot = df_con_periodos.groupby(['articulo', 'anio', 'mes'])['bultos'].sum().to_dict() if len(df_con_periodos) > 0 else {}
    art_totals = df_marca.groupby('articulo')['bultos'].sum().sort_values(ascending=False)
    art_ids = df_marca.groupby('articulo')['id_articulo'].first().to_dict()

    col_meses = [_periodo_label(int(a), int(m)) for a, m in periodos]
    headers = ['Cod', 'Articulo'] + col_meses + ['Total']

    if escribir_header:
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = s['FONT_HEADER']
            cell.fill = s['FILL_HEADER']
            cell.alignment = s['ALIGN_CENTER'] if c != 2 else Alignment()
        row += 1

    for articulo in art_totals.index:
        total = art_totals[articulo]
        cod = art_ids.get(articulo, '')
        ws.cell(row=row, column=1, value=cod).alignment = s['ALIGN_CENTER']
        ws.cell(row=row, column=2, value=articulo)
        for ci, (anio, mes) in enumerate(periodos):
            val = pivot.get((articulo, anio, mes))
            cell = ws.cell(row=row, column=3 + ci, value=val if val else None)
            cell.alignment = s['ALIGN_RIGHT']
            cell.number_format = '#,##0'
        cell_total = ws.cell(row=row, column=3 + len(periodos), value=total)
        cell_total.font = s['FONT_BOLD']
        cell_total.alignment = s['ALIGN_RIGHT']
        cell_total.number_format = '#,##0'
        row += 1

    # Fila subtotal
    ws.cell(row=row, column=2, value='Total').font = s['FONT_BOLD']
    mes_totals = df_con_periodos.groupby(['anio', 'mes'])['bultos'].sum().to_dict() if len(df_con_periodos) > 0 else {}
    for ci, (anio, mes) in enumerate(periodos):
        val = mes_totals.get((anio, mes))
        cell = ws.cell(row=row, column=3 + ci, value=val if val else None)
        cell.font = s['FONT_BOLD']
        cell.alignment = s['ALIGN_RIGHT']
        cell.number_format = '#,##0'
        cell.fill = s['FILL_SUBTOTAL']
    cell_gt = ws.cell(row=row, column=3 + len(periodos), value=df_marca['bultos'].sum())
    cell_gt.font = s['FONT_BOLD']
    cell_gt.alignment = s['ALIGN_RIGHT']
    cell_gt.number_format = '#,##0'
    cell_gt.fill = s['FILL_SUBTOTAL']
    row += 1

    return row


def _generar_excel_marca(id_cliente, generico, marca):
    """Genera Excel para una marca especifica."""
    from openpyxl import Workbook
    s = _get_excel_styles()

    df_all, periodos = _preparar_datos_excel(id_cliente)
    if len(df_all) == 0:
        return None

    df_marca = df_all[(df_all['generico'] == generico) & (df_all['marca'] == marca)]
    if len(df_marca) == 0:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = marca[:31]

    ws.cell(row=1, column=1, value=f'{generico} — {marca}').font = s['FONT_GENERICO']
    row = 3
    row = _escribir_tabla_marca(ws, row, df_marca, periodos)

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    n_cols = 3 + len(periodos)
    for c in range(3, n_cols + 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _generar_excel_completo(id_cliente):
    """Genera Excel con 3 hojas: Por Genérico, Por Marca, Detalle Artículos."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    s = _get_excel_styles()

    df_info = cargar_info_cliente(id_cliente)
    df_all, periodos = _preparar_datos_excel(id_cliente)
    if len(df_all) == 0:
        return None

    cliente_nombre = df_info.iloc[0]['razon_social'] if len(df_info) > 0 else f'Cliente {id_cliente}'
    col_meses = [_periodo_label(int(a), int(m)) for a, m in periodos]
    n_col_total = 2 + len(periodos)  # nombre + meses + total

    df_con_periodos = df_all.dropna(subset=['anio', 'mes'])

    wb = Workbook()

    # ===================== HOJA 1: POR GENÉRICO =====================
    ws_gen = wb.active
    ws_gen.title = 'Por Generico'
    ws_gen.cell(row=1, column=1, value=f'{cliente_nombre} [{id_cliente}]').font = s['FONT_CLIENTE']

    headers_gen = ['Generico'] + col_meses + ['Total']
    for c, h in enumerate(headers_gen, 1):
        cell = ws_gen.cell(row=3, column=c, value=h)
        cell.font = s['FONT_HEADER']
        cell.fill = s['FILL_HEADER']
        cell.alignment = s['ALIGN_CENTER'] if c != 1 else Alignment()

    gen_totals = df_all.groupby('generico')['bultos'].sum().sort_values(ascending=False)
    row = 4
    gran_total_meses = {(a, m): 0 for a, m in periodos}
    gran_total = 0

    for generico in gen_totals.index:
        df_gen_p = df_con_periodos[df_con_periodos['generico'] == generico]
        mes_totals = df_gen_p.groupby(['anio', 'mes'])['bultos'].sum().to_dict() if len(df_gen_p) > 0 else {}
        total = gen_totals[generico]

        ws_gen.cell(row=row, column=1, value=generico).font = s['FONT_BOLD']
        for ci, (anio, mes) in enumerate(periodos):
            val = mes_totals.get((anio, mes))
            cell = ws_gen.cell(row=row, column=2 + ci, value=val if val else None)
            cell.alignment = s['ALIGN_RIGHT']
            cell.number_format = '#,##0'
            if val:
                gran_total_meses[(anio, mes)] += val
        cell_t = ws_gen.cell(row=row, column=n_col_total, value=total)
        cell_t.font = s['FONT_BOLD']
        cell_t.alignment = s['ALIGN_RIGHT']
        cell_t.number_format = '#,##0'
        gran_total += total
        row += 1

    # Fila TOTAL
    ws_gen.cell(row=row, column=1, value='TOTAL').font = s['FONT_TOTAL']
    for c in range(1, n_col_total + 1):
        ws_gen.cell(row=row, column=c).fill = s['FILL_TOTAL']
    for ci, (anio, mes) in enumerate(periodos):
        val = gran_total_meses.get((anio, mes), 0)
        cell = ws_gen.cell(row=row, column=2 + ci, value=val if val else None)
        cell.font = s['FONT_BOLD']
        cell.alignment = s['ALIGN_RIGHT']
        cell.number_format = '#,##0'
        cell.fill = s['FILL_TOTAL']
    cell_gt = ws_gen.cell(row=row, column=n_col_total, value=gran_total)
    cell_gt.font = s['FONT_BOLD']
    cell_gt.alignment = s['ALIGN_RIGHT']
    cell_gt.number_format = '#,##0'
    cell_gt.fill = s['FILL_TOTAL']

    ws_gen.column_dimensions['A'].width = 25
    for c in range(2, n_col_total + 1):
        ws_gen.column_dimensions[ws_gen.cell(row=3, column=c).column_letter].width = 12

    # ===================== HOJA 2: POR MARCA =====================
    ws_marca = wb.create_sheet('Por Marca')
    ws_marca.cell(row=1, column=1, value=f'{cliente_nombre} [{id_cliente}]').font = s['FONT_CLIENTE']

    headers_marca = ['Generico', 'Marca'] + col_meses + ['Total']
    n_col_marca = 3 + len(periodos)  # generico + marca + meses + total
    for c, h in enumerate(headers_marca, 1):
        cell = ws_marca.cell(row=3, column=c, value=h)
        cell.font = s['FONT_HEADER']
        cell.fill = s['FILL_HEADER']
        cell.alignment = s['ALIGN_CENTER'] if c > 2 else Alignment()

    row = 4
    for generico in gen_totals.index:
        df_gen = df_all[df_all['generico'] == generico]
        marca_totals = df_gen.groupby('marca')['bultos'].sum().sort_values(ascending=False)
        gen_mes = {}

        for marca in marca_totals.index:
            df_m_p = df_con_periodos[(df_con_periodos['generico'] == generico) & (df_con_periodos['marca'] == marca)]
            mes_totals = df_m_p.groupby(['anio', 'mes'])['bultos'].sum().to_dict() if len(df_m_p) > 0 else {}
            total = marca_totals[marca]

            ws_marca.cell(row=row, column=1, value=generico)
            ws_marca.cell(row=row, column=2, value=marca)
            for ci, (anio, mes) in enumerate(periodos):
                val = mes_totals.get((anio, mes))
                cell = ws_marca.cell(row=row, column=3 + ci, value=val if val else None)
                cell.alignment = s['ALIGN_RIGHT']
                cell.number_format = '#,##0'
                if val:
                    gen_mes[(anio, mes)] = gen_mes.get((anio, mes), 0) + val
            cell_t = ws_marca.cell(row=row, column=n_col_marca, value=total)
            cell_t.font = s['FONT_BOLD']
            cell_t.alignment = s['ALIGN_RIGHT']
            cell_t.number_format = '#,##0'
            row += 1

        # Subtotal genérico
        ws_marca.cell(row=row, column=1, value=f'TOTAL {generico}').font = s['FONT_TOTAL']
        for c in range(1, n_col_marca + 1):
            ws_marca.cell(row=row, column=c).fill = s['FILL_GENERICO']
        for ci, (anio, mes) in enumerate(periodos):
            val = gen_mes.get((anio, mes), 0)
            cell = ws_marca.cell(row=row, column=3 + ci, value=val if val else None)
            cell.font = s['FONT_BOLD']
            cell.alignment = s['ALIGN_RIGHT']
            cell.number_format = '#,##0'
            cell.fill = s['FILL_GENERICO']
        cell_gt = ws_marca.cell(row=row, column=n_col_marca, value=gen_totals[generico])
        cell_gt.font = s['FONT_BOLD']
        cell_gt.alignment = s['ALIGN_RIGHT']
        cell_gt.number_format = '#,##0'
        cell_gt.fill = s['FILL_GENERICO']
        row += 1

    ws_marca.column_dimensions['A'].width = 20
    ws_marca.column_dimensions['B'].width = 25
    for c in range(3, n_col_marca + 1):
        ws_marca.column_dimensions[ws_marca.cell(row=3, column=c).column_letter].width = 12

    # ===================== HOJA 3: DETALLE ARTÍCULOS =====================
    ws_det = wb.create_sheet('Detalle Articulos')
    ws_det.cell(row=1, column=1, value=f'{cliente_nombre} [{id_cliente}]').font = s['FONT_CLIENTE']
    row = 3

    n_cols_det = 3 + len(periodos)  # cod + articulo + meses + total
    for generico in gen_totals.index:
        df_gen = df_all[df_all['generico'] == generico]

        cell_gen = ws_det.cell(row=row, column=1, value=generico)
        cell_gen.font = s['FONT_GENERICO']
        for c in range(1, n_cols_det + 1):
            ws_det.cell(row=row, column=c).fill = s['FILL_GENERICO']
        row += 1

        marca_totals = df_gen.groupby('marca')['bultos'].sum().sort_values(ascending=False)
        gen_mes_totals = {(a, m): 0 for a, m in periodos}
        gen_gran_total = 0

        for marca in marca_totals.index:
            df_marca = df_gen[df_gen['marca'] == marca]

            cell_m = ws_det.cell(row=row, column=1, value=f'  {marca}')
            cell_m.font = s['FONT_MARCA']
            for c in range(1, n_cols_det + 1):
                ws_det.cell(row=row, column=c).fill = s['FILL_MARCA']
            row += 1

            row = _escribir_tabla_marca(ws_det, row, df_marca, periodos)

            df_con_p = df_marca.dropna(subset=['anio', 'mes'])
            if len(df_con_p) > 0:
                for (a, m), val in df_con_p.groupby(['anio', 'mes'])['bultos'].sum().items():
                    gen_mes_totals[(a, m)] = gen_mes_totals.get((a, m), 0) + val
            gen_gran_total += df_marca['bultos'].sum()
            row += 1

        ws_det.cell(row=row, column=1, value=f'TOTAL {generico}').font = s['FONT_TOTAL']
        for ci, (anio, mes) in enumerate(periodos):
            val = gen_mes_totals.get((anio, mes), 0)
            cell = ws_det.cell(row=row, column=3 + ci, value=val if val else None)
            cell.font = s['FONT_BOLD']
            cell.alignment = s['ALIGN_RIGHT']
            cell.number_format = '#,##0'
            cell.fill = s['FILL_TOTAL']
        cell_gt = ws_det.cell(row=row, column=n_cols_det, value=gen_gran_total)
        cell_gt.font = s['FONT_BOLD']
        cell_gt.alignment = s['ALIGN_RIGHT']
        cell_gt.number_format = '#,##0'
        cell_gt.fill = s['FILL_TOTAL']
        row += 2

    ws_det.column_dimensions['A'].width = 10
    ws_det.column_dimensions['B'].width = 35
    for c in range(3, n_cols_det + 1):
        ws_det.column_dimensions[ws_det.cell(row=1, column=c).column_letter].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# =============================================================================
# CALLBACKS EXCEL
# =============================================================================

# Event delegation para botones Excel por marca.
# Se ejecuta una vez al cargar la pagina de detalle de cliente.
clientside_callback(
    """
    function(storeData) {
        if (window._clienteClickSetup) return dash_clientside.no_update;
        window._clienteClickSetup = true;

        document.addEventListener('click', function(e) {
            var excelBtn = e.target.closest('.btn-excel-marca');
            if (excelBtn) {
                var gen = excelBtn.getAttribute('data-generico');
                var marca = excelBtn.getAttribute('data-marca');
                dash_clientside.set_props('excel-marca-trigger', {
                    data: gen + '||' + marca + '||' + Date.now()
                });
            }
        });
        return dash_clientside.no_update;
    }
    """,
    Output('excel-marca-trigger', 'data'),
    Input('cliente-id-store', 'data'),
    prevent_initial_call=False
)


@callback(
    Output('download-excel-marca', 'data'),
    Input('excel-marca-trigger', 'data'),
    State('cliente-id-store', 'data'),
    prevent_initial_call=True
)
def exportar_marca_excel(trigger_value, store_data):
    """Exporta a Excel la tabla de una marca especifica."""
    if not trigger_value or '||' not in str(trigger_value) or not store_data:
        return no_update

    parts = str(trigger_value).split('||')
    if len(parts) < 2:
        return no_update

    generico, marca = parts[0], parts[1]
    id_cliente = store_data['id_cliente']

    excel_bytes = _generar_excel_marca(id_cliente, generico, marca)
    if excel_bytes is None:
        return no_update

    filename = f"{generico}_{marca}.xlsx".replace(' ', '_')
    return dcc.send_bytes(excel_bytes, filename)


@callback(
    Output('download-excel-completo', 'data'),
    Input('btn-excel-completo', 'n_clicks'),
    State('cliente-id-store', 'data'),
    prevent_initial_call=True
)
def exportar_completo_excel(n_clicks, store_data):
    """Exporta a Excel todo el arbol generico->marca->articulo con subtotales."""
    if not n_clicks or not store_data:
        return no_update

    id_cliente = store_data['id_cliente']

    excel_bytes = _generar_excel_completo(id_cliente)
    if excel_bytes is None:
        return no_update

    filename = f"cliente_{id_cliente}_completo.xlsx"
    return dcc.send_bytes(excel_bytes, filename)
